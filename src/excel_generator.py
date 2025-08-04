import pandas as pd
from pathlib import Path
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# Путь для сохранения отчетов
REPORTS_DIR = Path(__file__).parent.parent / "reports"

def _clean_reports_directory():
    """
    Очищает папку reports, удаляя все файлы для экономии места на сервере.
    """
    if REPORTS_DIR.exists():
        try:
            # Удаляем все файлы в папке
            for file_path in REPORTS_DIR.iterdir():
                if file_path.is_file():
                    file_path.unlink()
                    print(f"Удален файл: {file_path.name}")
                elif file_path.is_dir():
                    import shutil
                    shutil.rmtree(file_path)
                    print(f"Удалена папка: {file_path.name}")
            print(f"Папка reports очищена успешно")
        except Exception as e:
            print(f"Ошибка при очистке папки reports: {e}")
    else:
        print("Папка reports не существует, создание не требуется")

def _is_valid_url(url):
    """Проверяет, является ли строка валидным URL."""
    try:
        from urllib.parse import urlparse
        result = urlparse(url)
        return all([result.scheme, result.netloc])
    except (ValueError, AttributeError):
        return False

def _format_price(price: float) -> str:
    """Форматирует цену в удобный для чтения формат (тыс. или млн. руб)."""
    if not isinstance(price, (int, float)) or price == 0:
        return "0 руб."
    
    price = float(price)
    
    if price >= 1_000_000:
        price_in_millions = round(price / 1_000_000)
        return f"{price_in_millions} млн. руб."
    elif price >= 1_000:
        price_in_thousands = round(price / 1_000)
        return f"{price_in_thousands} тыс. руб."
    else:
        return f"{int(price)} руб."

def create_excel_report(listings: list[dict], city: str = "") -> Path | None:
    """
    Создает Excel-отчет без изображений.
    city: название города (для имени файла)
    """
    if not listings:
        print("Нет данных для создания Excel-отчета.")
        return None

    REPORTS_DIR.mkdir(exist_ok=True)
    
    df = pd.DataFrame(listings)
    
    # Применяем форматирование к колонкам с ценами
    if 'price_per_sqm' in df.columns:
        df['price_per_sqm'] = df['price_per_sqm'].apply(_format_price)
    if 'price' in df.columns:
        df['price'] = df['price'].apply(_format_price)

    # Определяем правильную единицу измерения площади в зависимости от категории
    def get_area_column_name(category):
        category_lower = str(category).lower()
        if 'земельные' in category_lower or 'земельный' in category_lower:
            return 'Площадь, сотки'
        else:
            return 'Площадь, кв.м.'
    
    # Применяем правильное название колонки для площади
    if 'category_name' in df.columns:
        # Определяем преобладающую категорию для правильного названия колонки
        categories = df['category_name'].dropna().unique()
        area_column_name = 'Площадь, кв.м.'  # по умолчанию
        for category in categories:
            category_lower = str(category).lower()
            if 'земельные' in category_lower or 'земельный' in category_lower:
                area_column_name = 'Площадь, сотки'
                break
    else:
        area_column_name = 'Площадь, кв.м.'
    
    report_df = df.rename(columns={
        'address': 'Адрес', 'area': area_column_name, 
        'price_per_sqm': 'Цена за кв.м.', 'price': 'Итоговая цена',
        'description': 'Описание', 'url': 'Ссылка на объявление',
        'category_name': 'Категория', 'category_color': 'Цвет категории'
    })
    
    # Оставляем только нужные колонки (убираем колонку с изображениями)
    columns_to_include = ['Адрес', 'Категория', area_column_name, 'Цена за кв.м.', 'Итоговая цена', 'Описание', 'Ссылка на объявление']
    available_columns = [col for col in columns_to_include if col in report_df.columns]
    report_df = report_df[available_columns]
    
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    city_part = f"_{city.lower().replace(' ', '_')}" if city else ""
    file_path = REPORTS_DIR / f"realty_report{city_part}_{timestamp}.xlsx"
    
    # 1. Сохраняем текстовые данные
    report_df.to_excel(file_path, index=False, engine='openpyxl')

    # 2. Открываем файл с помощью openpyxl для настройки стилей
    wb = load_workbook(file_path)
    ws = wb.active

    # --- Поиск колонок и настройка стилей ---
    header = [cell.value for cell in ws[1]]
    try:
        url_col_letter = get_column_letter(header.index('Ссылка на объявление') + 1)
    except ValueError:
        print("Критическая ошибка: не найдена колонка 'Ссылка на объявление'.")
        return file_path

    hyperlink_font = Font(color="0000FF", underline="single")

    # --- Настройка ширины колонок ---
    for i, col_name in enumerate(header):
        # Простой автофит
        max_len = report_df[col_name].astype(str).map(len).max() or len(col_name)
        ws.column_dimensions[get_column_letter(i + 1)].width = max(len(col_name), max_len) + 2

    # --- Итерация по строкам для настройки ссылок ---
    for row_num in range(2, ws.max_row + 1):
        # --- Гиперссылка на объявление ---
        url_cell = ws[f'{url_col_letter}{row_num}']
        if url_cell.value and _is_valid_url(url_cell.value):
            url_cell.hyperlink = url_cell.value
            url_cell.value = "Перейти"
            url_cell.font = hyperlink_font

    # 3. Сохраняем итоговый файл
    wb.save(file_path)

    print(f"Excel-отчет успешно создан: {file_path}")
    return file_path

if __name__ == '__main__':
    print("Создание тестового Excel-отчета...")
    
    # Создаем тестовые данные
    test_listings = []
    for i in range(1, 11):  # 10 записей для теста
        test_listings.append({
            'address': f'Тестовый проспект, {i}', 
            'area': 75.5 + i, 
            'price_per_sqm': 158940.4 + i * 1000, 
            'price': 12000000.0 + i * 100000, 
            'description': f'Отличное помещение в центре города #{i}.', 
            'url': f'https://example.com/{i}',
            'title': f'Помещение {75.5 + i}м', 
            'category_name': 'Коммерческая недвижимость', 
            'category_color': 'red'
        })
    
    report_path = create_excel_report(test_listings, city="Тестовый город")
    if report_path:
        print(f"Тестовый отчет создан: {report_path}")
    else:
        print("Ошибка при создании тестового отчета") 